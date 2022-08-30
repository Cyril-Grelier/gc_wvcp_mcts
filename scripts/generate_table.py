"""
Generate a xls file containing results of given methods from csv files.

Follow the main function to manage the files and parameters

Gap compute the difference on the mean of two methods and give the p-value
with a ttest.
"""
import statistics
from glob import glob
import re
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from scipy import stats

# Only the P_VALUE_1 will be take in account to count the number of time a method is better.
# The two other will be colored with colors COLOR_GAP{1,2}_{1,2,3}.
P_VALUE_1 = 0.001
P_VALUE_2 = 0.05
P_VALUE_3 = 0.1
COLOR_GAP1_1 = "488f31"
COLOR_GAP1_2 = "76a263"
COLOR_GAP1_3 = "9fb494"
COLOR_GAP2_1 = "de425b"
COLOR_GAP2_2 = "dd757d"
COLOR_GAP2_3 = "d69fa1"

# When a best score is equal to the best known score : red
COLOR_BEST = "ff0000"
# When a best score is proven optimal : blue
COLOR_OPTIMAL = "0000ff"
# When a best score is better than the best known score : green
COLOR_NEW_BEST = "00ff00"


def main():
    """
    Choose the methods, gaps and instance and create the xls file
    """
    # Add method name, repertory of data and short name of the method
    methods: list[Method] = []

    lss: tuple[str, str] = [
        ("random", "Random"),
        ("constrained", "Greedy-Random"),
        ("deterministic", "Greedy"),
    ]
    for l_s in lss:
        methods.append(
            Method(
                f"{l_s[1]}",
                f"outputs/greedy_only_all/{l_s[0]}/",
                f"{l_s[1]}",
            )
        )
        methods.append(
            Method(
                f"MCTS+{l_s[1]}",
                f"outputs/mcts_3_greedy/{l_s[0]}/",
                f"MCTS+{l_s[1]}",
            )
        )
    lss = [
        ("afisa_original", "AFISA"),
        ("tabu_weight", "TabuWeight"),
        ("redls", "RedLS"),
        ("ilsts", "ILSTS"),
    ]
    for l_s in lss:
        methods.append(
            Method(
                f"{l_s[1]}",
                f"outputs/ls_all_1h/{l_s[0]}/",
                f"{l_s[1]}",
            )
        )
        methods.append(
            Method(
                f"MCTS+{l_s[1]}",
                f"outputs/mcts_ls_all_1h/{l_s[0]}/",
                f"MCTS+{l_s[1]}",
            )
        )

    problem = "wvcp"

    # Choose the set of instances
    # instances_set = ("pxx", "pxx")
    # instances_set = ("rxx", "rxx")
    # instances_set = ("DIMACS_optimal", "dimacs_o")
    # instances_set = ("DIMACS_non_optimal", "dimacs_no")
    instances_set = ("instance_list_wvcp", "all")

    output_file = f"xlsx_files/greedy_vs_ls_vs_mcts_{instances_set[1]}.xlsx"

    # Choose the method to compare with ttest (just need the method name and short name)
    gaps = [
        (methods[i], methods[j])
        for i in range(len(methods))
        for j in range(i + 1, len(methods))
    ]

    with open(f"instances/{instances_set[0]}.txt", "r", encoding="utf8") as file:
        instances = [i[:-1] for i in file.readlines()]

    table = Table(methods=methods, instances=instances, gaps=gaps, problem=problem)
    table.to_xls(output_file)
    print(output_file)


@dataclass
class TimeScore:
    """Store time and score"""

    time: int
    score: int

    def __repr__(self) -> str:
        return f"t:{self.time},s:{self.score}"


@dataclass
class FinalResults:
    """Store line of best results for a method"""

    best: int
    average: float
    time: float
    nb_best: int
    nb_total: int

    def to_sheet_line(self) -> list:
        """
        Convert self to a list for the sheet,
        edit length_line and sheet_line_content if you edit this function
        """
        return [self.best, self.average, self.time, f"{self.nb_best}/{self.nb_total}"]

    @staticmethod
    def length_line():
        """Return number of columns for a method"""
        return len(FinalResults.sheet_line_content())

    @staticmethod
    def sheet_line_content():
        """Header of method"""
        return ["best", "avg", "time", "#"]

    def __repr__(self) -> str:
        return f"b:{self.best},a:{self.average},t:{self.time},{self.nb_best}/{self.nb_total}"


@dataclass
class Gap:
    """Store the gap and p-value between 2 methods"""

    difference: float
    p_value: float

    def to_sheet_line(self):
        """Convert self to a list for the sheet"""
        return [self.difference, self.p_value]

    @staticmethod
    def length_line():
        """Return number of columns for a gap"""
        return 2

    @staticmethod
    def sheet_line_content():
        """Header of gap"""
        return ["diff", "pvalue"]

    def __repr__(self) -> str:
        return f"d:{self.difference},p:{self.p_value}"


@dataclass
class Method:
    """Store name, repertory and short name of a method/algorithm"""

    name: str
    repertory: str
    short_name: str

    def __repr__(self) -> str:
        return self.name

    def __hash__(self) -> int:
        return hash(str(self))


@dataclass
class Instance:
    """Store the results of all given method on one instance"""

    name: str
    nb_vertices: int
    nb_edges: int
    best_known_score: int
    optimal: bool
    methods: list[Method]
    method_optimal: dict[str, bool]
    raw_results: dict[str, list[TimeScore]]
    final_results: dict[str, FinalResults]
    gaps: dict[tuple[str, str], Gap]

    def compute_final_results(self) -> None:
        """Compute results from raw data"""
        for method in self.methods:
            data: list[TimeScore] = self.raw_results[method.name]
            try:
                best_score = min([line.score for line in data])
            except TypeError:
                best_score = 999999
            except Exception as exception:
                print("error :", self.name, method.__dict__)
                raise exception
            self.final_results[method.name] = FinalResults(
                best=best_score,
                average=round(statistics.mean([line.score for line in data]), 1)
                if best_score != 999999
                else 999999,
                time=round(
                    statistics.mean(
                        [line.time for line in data if line.score == best_score]
                    ),
                    1,
                )
                if best_score != 999999
                else 999999,
                nb_best=len([True for line in data if line.score == best_score])
                if best_score != 999999
                else 0,
                nb_total=len(data) if best_score != 999999 else 0,
            )

    def compute_gap(self, methods_names: tuple[Method, Method]):
        """Compute gap"""
        try:
            s_1 = [line.score for line in self.raw_results[methods_names[0].name]]
        except TypeError:
            s_1 = []
        try:
            s_2 = [line.score for line in self.raw_results[methods_names[1].name]]
        except TypeError:
            s_2 = []
        _, p_value = stats.ttest_ind(s_1, s_2)
        self.gaps[methods_names] = Gap(
            # difference=round(statistics.mean(s1) - statistics.mean(s2), 2),
            difference=round(
                self.final_results[methods_names[0].name].average
                - self.final_results[methods_names[1].name].average,
                1,
            ),
            p_value=round(p_value, 3),
        )

    def to_sheet_line(self):
        """
        Convert self to a list for the sheet,
        edit length_line and sheet_line_content if you edit this function
        """
        line = [
            self.name,
            self.nb_vertices,
            self.nb_edges,
            self.best_known_score,
            self.optimal,
        ]
        for method in self.methods:
            line.extend(self.final_results[method.name].to_sheet_line())
        for _, gap in self.gaps.items():
            line.extend(gap.to_sheet_line())
        return line

    @staticmethod
    def length_line():
        """Return number of columns for a instance"""
        return 5

    @staticmethod
    def sheet_line_content():
        """Header of instance"""
        return ["instance", "|V|", "|E|", "BKS", "optim"]
        # return ["instance", "|V|", "BKS", "optim"]


def get_nb_vertices_edges(instance: str) -> tuple[int, int]:
    """Read the number of vertices

    Args:
        instance (str): instance name

    Raises:
        Exception: Instance not found

    Returns:
        tuple[int, int]: number of vertices and number of edges
    """
    with open("instances/instance_info.txt", "r", encoding="utf8") as file:
        for line in file.readlines():
            instance_, nb_vertices, nb_edges = line[:-1].split(",")
            if instance_ == instance:
                return nb_vertices, nb_edges
    raise Exception(f"instance {instance} not found in instances/instance_info.txt")


def get_best_known_score(instance: str, problem: str) -> tuple[int, bool]:
    """Read best know score

    Args:
        instance (str): instance name
        problem (str): type of problem (gcp, wvcp)

    Raises:
        Exception: Instance not found

    Returns:
        tuple[int, bool]: best known score and optimality
    """
    file = f"instances/best_scores_{problem}.txt"
    with open(file, "r", encoding="utf8") as file:
        for line in file.readlines():
            instance_, score, optimal = line[:-1].split(" ")
            if instance_ == instance:
                return int(score), optimal == "*"
    raise Exception(f"instance {instance} not found in {file}")


def load_instance(instance: str, methods: list[Method], problem: str) -> Instance:
    """Load an instance by its name and each method listed

    Args:
        instance (str): instance
        methods (list[Method]): the methods
        problem (str): type of problem (gcp, wvcp)

    Returns:
        Instance: the loaded instance
    """
    nb_vertices, nb_edges = get_nb_vertices_edges(instance)
    best_known_score, optimal = get_best_known_score(instance, problem)
    return Instance(
        name=instance,
        nb_vertices=nb_vertices,
        nb_edges=nb_edges,
        best_known_score=best_known_score,
        optimal=optimal,
        methods=methods,
        method_optimal={},
        raw_results={},
        final_results={},
        gaps={},
    )


class Table:
    """Representation of the data table"""

    def __init__(
        self,
        methods: list[Method],
        instances: list[str],
        gaps: list[tuple[Method, Method]],
        problem: str,
    ) -> None:
        self.methods: list[Method] = methods
        self.instances: list[Instance] = [
            load_instance(instance, methods, problem) for instance in instances
        ]
        self.gaps: list[tuple[Method, Method]] = gaps
        self.problem: str = problem
        for instance in self.instances:
            print(instance.name)
            for method in self.methods:
                # get the last line of the file and convert time and score to int
                instance.raw_results[method.name] = [
                    TimeScore(
                        *map(
                            int,
                            pd.read_csv(file_name)[["time", "score"]]
                            .iloc[-1]
                            .to_list(),
                        )
                    )
                    for file_name in sorted(
                        glob(f"{method.repertory}{instance.name}_[0-9]*.csv"),
                        key=lambda f: int(re.sub(r"\D", "", f)),
                    )
                ]
                instance.method_optimal[method.name] = False
                if not instance.raw_results[method.name]:
                    instance.raw_results[method.name] = None
                # for mcts only, check for optimality
                try:
                    for file_name in sorted(
                        glob(f"{method.repertory}{instance.name}_[0-9]*.csv"),
                        key=lambda f: int(re.sub(r"\D", "", f)),
                    ):
                        nb_total_node, nb_current_node = (
                            pd.read_csv(file_name)[
                                [
                                    "nb total node",
                                    "nb current node",
                                ]
                            ]
                            .iloc[-1]
                            .to_list()
                        )
                        if nb_total_node > 1 and nb_current_node <= 1:
                            instance.method_optimal[method.name] = True
                except KeyError:
                    pass

        print("raw results loaded")
        print("compute final results")
        for instance in self.instances:
            instance.compute_final_results()
        print("compute final results done")
        print("compute gap")
        for instance in self.instances:
            for gap in self.gaps:
                instance.compute_gap(gap)
        print("compute gap done")

    def __repr__(self) -> str:
        return "\n".join([str(instance) for instance in self.instances])

    def to_xls(self, file_name: str):
        """Convert the table to xls file"""
        print("conversion to xlsx")
        workbook = Workbook()
        sheet = workbook.active
        # first row
        header1 = Instance.sheet_line_content()
        firsts_cols = Instance.length_line()
        nb_col_met = FinalResults.length_line()
        nb_col_gap = Gap.length_line()

        for method in self.methods:
            header1.extend([method.name] * nb_col_met)
        for gap in self.gaps:
            header1.extend([f"gap {gap[0].short_name}-{gap[1].short_name}", ""])
        sheet.append(header1)
        # second row
        header2 = [""] * firsts_cols
        for method in self.methods:
            header2.extend(FinalResults.sheet_line_content())
        for _ in self.gaps:
            header2.extend(Gap.sheet_line_content())
        sheet.append(header2)

        # merge
        for i in range(firsts_cols):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )
        for i in range(len(self.methods)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=firsts_cols + 1 + nb_col_met * i,
                end_column=firsts_cols + nb_col_met * (i + 1),
            )
        start_gap = firsts_cols + nb_col_met * len(self.methods) + 1
        for i in range(len(self.gaps)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=start_gap + nb_col_gap * i,
                end_column=start_gap + nb_col_gap * i + 1,
            )

        # add data
        for instance in self.instances:
            sheet.append(instance.to_sheet_line())

        # add color to best scores
        nb_best = {method.name: 0 for method in self.methods}
        nb_optimal = {method.name: 0 for method in self.methods}
        for row, instance in enumerate(self.instances, start=3):
            bks = instance.best_known_score
            for col, method in enumerate(self.methods):
                cell = sheet.cell(row, firsts_cols + 1 + nb_col_met * col)
                if int(cell.value) == int(bks):
                    # red for method == best know score
                    nb_best[method.name] += 1
                    cell.font = Font(bold=True, color=COLOR_BEST)
                elif int(cell.value) < int(bks):
                    # green for method better than best know score
                    nb_best[method.name] += 1
                    cell.font = Font(bold=True, color=COLOR_NEW_BEST)
                if instance.method_optimal[method.name]:
                    # blue for proven optimal score
                    nb_optimal[method.name] += 1
                    # cell.fill = PatternFill(bgColor="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True, color=COLOR_OPTIMAL)

        # add color to gaps
        begin_gap = firsts_cols + 1 + nb_col_met * len(self.methods)
        for row in range(3, len(self.instances) + 3):
            for col in range(len(self.gaps)):
                diff_cell = sheet.cell(row, begin_gap + col * 2)
                p_val_cell = sheet.cell(row, begin_gap + col * 2 + 1)
                color = "808080"  # gray
                if float(diff_cell.value) > 0:
                    # 1 better
                    if float(p_val_cell.value) <= P_VALUE_1:
                        color = COLOR_GAP1_1
                    elif float(p_val_cell.value) <= P_VALUE_2:
                        color = COLOR_GAP1_2
                    elif float(p_val_cell.value) <= P_VALUE_3:
                        color = COLOR_GAP1_3
                elif float(diff_cell.value) < 0:
                    # 2 better
                    if float(p_val_cell.value) <= P_VALUE_1:
                        color = COLOR_GAP2_1
                    elif float(p_val_cell.value) <= P_VALUE_2:
                        color = COLOR_GAP2_2
                    elif float(p_val_cell.value) <= P_VALUE_3:
                        color = COLOR_GAP2_3
                diff_cell.font = Font(bold=True, color=color)
                p_val_cell.font = Font(bold=True, color=color)
        # footer
        last_line = [""] * firsts_cols
        last_line[0] = "nb total best"
        for i, method in enumerate(self.methods):
            last_line.extend(
                [f"{nb_best[method.name]}/{len(self.instances)}"] * nb_col_met
            )
        nb_positive_gap = {gap: 0 for gap in self.gaps}
        nb_negative_gap = {gap: 0 for gap in self.gaps}
        for instance in self.instances:
            for gap_name, gap in instance.gaps.items():
                if gap.difference < 0 and gap.p_value <= P_VALUE_1:  # p-value
                    nb_negative_gap[gap_name] += 1
                if gap.difference > 0 and gap.p_value <= P_VALUE_1:
                    nb_positive_gap[gap_name] += 1
        for gap in self.gaps:
            last_line.extend([f"{gap[1].short_name} better", nb_positive_gap[gap]])
        sheet.append(last_line)

        sheet.merge_cells(
            start_row=len(self.instances) + 3,
            end_row=len(self.instances) + 3,
            start_column=1,
            end_column=firsts_cols,
        )

        for i, _ in enumerate(self.methods):
            sheet.merge_cells(
                start_row=len(self.instances) + 3,
                end_row=len(self.instances) + 3,
                start_column=firsts_cols + 1 + nb_col_met * i,
                end_column=firsts_cols + nb_col_met * (i + 1),
            )

        last_line = [""] * firsts_cols
        last_line[0] = "nb total optim"
        for i, method in enumerate(self.methods):
            last_line.extend(
                [f"{nb_optimal[method.name]}/{len(self.instances)}"] * nb_col_met
            )
        for gap in self.gaps:
            last_line.extend([f"{gap[0].short_name} better", nb_negative_gap[gap]])
        sheet.append(last_line)
        sheet.merge_cells(
            start_row=len(self.instances) + 4,
            end_row=len(self.instances) + 4,
            start_column=1,
            end_column=firsts_cols,
        )

        for i, _ in enumerate(self.methods):
            sheet.merge_cells(
                start_row=len(self.instances) + 4,
                end_row=len(self.instances) + 4,
                start_column=firsts_cols + 1 + nb_col_met * i,
                end_column=firsts_cols + nb_col_met * (i + 1),
            )

        # add color to gaps names
        for i, row in enumerate(
            range(len(self.instances) + 3, len(self.instances) + 5)
        ):
            for col in range(len(self.gaps)):
                diff_cell = sheet.cell(row, begin_gap + col * 2)
                p_val_cell = sheet.cell(row, begin_gap + col * 2 + 1)
                if int(p_val_cell.value) != 0 and i == 0:
                    diff_cell.font = Font(bold=True, color=COLOR_GAP1_1)
                    p_val_cell.font = Font(bold=True, color=COLOR_GAP1_1)
                elif int(p_val_cell.value) != 0 and i == 1:
                    diff_cell.font = Font(bold=True, color=COLOR_GAP2_1)
                    p_val_cell.font = Font(bold=True, color=COLOR_GAP2_1)

        nb_positive = {method: 0 for method in self.methods}
        nb_negative = {method: 0 for method in self.methods}

        for method1_2, value in nb_positive_gap.items():
            method1, method2 = method1_2
            nb_positive[method2] += value
            nb_negative[method1] += value

        for method1_2, value in nb_negative_gap.items():
            method1, method2 = method1_2
            nb_positive[method1] += value
            nb_negative[method2] += value

        last_line = [""] * firsts_cols
        last_line[0] = "nb times significantly better"
        for i, method in enumerate(self.methods):
            last_line.extend([f"{nb_positive[method]}"] * nb_col_met)
        sheet.append(last_line)
        sheet.merge_cells(
            start_row=len(self.instances) + 5,
            end_row=len(self.instances) + 5,
            start_column=1,
            end_column=firsts_cols,
        )
        for i, _ in enumerate(self.methods):
            sheet.merge_cells(
                start_row=len(self.instances) + 5,
                end_row=len(self.instances) + 5,
                start_column=firsts_cols + 1 + nb_col_met * i,
                end_column=firsts_cols + nb_col_met * (i + 1),
            )

        last_line = [""] * firsts_cols
        last_line[0] = "nb times significantly worse"
        for i, method in enumerate(self.methods):
            last_line.extend([f"{nb_negative[method]}"] * nb_col_met)
        sheet.append(last_line)
        sheet.merge_cells(
            start_row=len(self.instances) + 6,
            end_row=len(self.instances) + 6,
            start_column=1,
            end_column=firsts_cols,
        )
        for i, _ in enumerate(self.methods):
            sheet.merge_cells(
                start_row=len(self.instances) + 6,
                end_row=len(self.instances) + 6,
                start_column=firsts_cols + 1 + nb_col_met * i,
                end_column=firsts_cols + nb_col_met * (i + 1),
            )

        last_line = [""] * firsts_cols
        last_line[0] = "ratio better/worse"
        for i, method in enumerate(self.methods):
            try:
                nb_to_print = f"{nb_positive[method] / nb_negative[method]:.2f}"
            except ZeroDivisionError:
                nb_to_print = "-"
            last_line.extend([nb_to_print] * nb_col_met)

        sheet.append(last_line)
        sheet.merge_cells(
            start_row=len(self.instances) + 7,
            end_row=len(self.instances) + 7,
            start_column=1,
            end_column=firsts_cols,
        )
        for i, _ in enumerate(self.methods):
            sheet.merge_cells(
                start_row=len(self.instances) + 7,
                end_row=len(self.instances) + 7,
                start_column=firsts_cols + 1 + nb_col_met * i,
                end_column=firsts_cols + nb_col_met * (i + 1),
            )

        # Set alignment
        for col in sheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["F3"]

        # Save
        workbook.save(file_name)


if __name__ == "__main__":
    main()
